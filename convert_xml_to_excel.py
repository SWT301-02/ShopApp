import pandas as pd
import xml.etree.ElementTree as ET
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Function to convert XML to DataFrame
def xml_to_dataframe(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    all_records = []
    for testcase in root.findall('testcase'):
        record = {}
        record['name'] = testcase.get('name')
        record['classname'] = testcase.get('classname')
        record['time'] = testcase.get('time')
        failure = testcase.find('failure')
        if failure is not None:
            record['failure'] = failure.text
        else:
            record['failure'] = None
        all_records.append(record)
    return pd.DataFrame(all_records)

# Convert all XML files in the target/surefire-reports directory
xml_files = glob.glob('target/surefire-reports/*.xml')
all_dfs = [xml_to_dataframe(xml_file) for xml_file in xml_files]
final_df = pd.concat(all_dfs, ignore_index=True)

# Save to Excel
excel_path = 'target/surefire-reports/test-report.xlsx'
final_df.to_excel(excel_path, index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(excel_path)
ws = wb.active

# Apply formatting
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center")

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save(excel_path)
